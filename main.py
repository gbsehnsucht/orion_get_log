from tkinter import *
from tkinter import ttk
from openpyxl import Workbook
from datetime import datetime
import pyodbc
import os


drivers_list = pyodbc.drivers()

query_columns = 'select column_name from information_schema.columns where table_name = '
query_data = f'select * from '

res_dir = '.\\results'


def get_data(conn):

    events_columns = conn.execute(query_columns + "'Events';").fetchall()
    events_list = [tuple(column_name[0] for column_name in events_columns)]
    events_list.extend(conn.execute(query_data + "Events;").fetchall())

    logs_columns = conn.execute(query_columns + "'pLogData';").fetchall()
    logs_list = [tuple(column_name[0] for column_name in logs_columns)]
    logs_list.extend(conn.execute(query_data + "pLogData order by TimeVal;").fetchall())

    alarm_columns = conn.execute(query_columns + "'m_alarm';").fetchall()
    alarms_list = [tuple(column_name[0] for column_name in alarm_columns)]
    alarms_list.extend(conn.execute(query_data + "m_alarm order by Time0;").fetchall())

    return [events_list, logs_list, alarms_list]


def conn_error_handle(exception):
    e_code, e_text = exception.args[0], exception.args[1]
    if e_code == 'HY000' or e_code == 'HYT00':
        return 'Неверно выбран драйвер!'
    if e_code == '08001':
        return 'Сервер не найден или недоступен!'
    if e_code == '28000' and 'Не удается открыть базу данных' in e_text:
        return 'Не найдена база данных!'
    if e_code == '28000':
        return 'Неверные UID/PWD!'


def frame_message_place(text):
    frame_message.place(relx=0.13, rely=0.76, height=40, width=280)
    lbl_message.configure(text=text)


def get_connection(driver, server, db, uid, pwd):
    connection_string = (
        f"Driver={driver};"
        f"Server={server};"
        f"Database={db};"
        f"UID={uid};"
        f"PWD={pwd};"
    )
    conn = pyodbc.connect(connection_string)
    return conn


def connect_to_sql():
    driver = combobox_driver.get()
    server = entry_server.get()
    db = entry_db.get()
    uid = entry_uid.get()
    pwd = entry_pwd.get()
    if '' in [driver, server, db, uid, pwd]:
        frame_message_place('Введены неполные данные!')
    else:
        try:
            return get_connection(driver, server, db, uid, pwd)
        except Exception as e:
            conn_error = conn_error_handle(e)
            frame_message_place(conn_error)


def create_xls_files(sql_log_data):
    wb = Workbook()
    sheet_names = ['events', 'logs', 'alarms']
    for section in range(len(sheet_names)):
        sheet_section = wb.create_sheet(sheet_names[section])
        sql_section = sql_log_data[section]
        for i in range(len(sql_section)):
            for j in range(len(sql_section[i])):
                sheet_section.cell(row=i + 1, column=j + 1).value = sql_section[i][j]
    del wb['Sheet']
    wb.save(f"{res_dir}\\{datetime.now().strftime('%Y-%m-%d %H_%M_%S')}.xlsx")


def create_res_dir():
    if not os.path.exists(res_dir):
        os.mkdir(res_dir)


def btn_connect_func():
    conn = connect_to_sql()
    if conn:
        create_res_dir()
        try:
            create_xls_files(get_data(conn))
            frame_message_place('Логи успешно выгружены!')
        except Exception as e:
            with open(f'{res_dir}\\fail.txt', 'w') as fail_txt:
                fail_txt.write(str(e))
                frame_message_place('Произошла ошибка! \n (смотри fail.txt)')


window = Tk()
window.title('orion_get_log')
window.minsize(width=400, height=300)
window.geometry('400x300')


lbl_driver = Label(text='Driver:')
lbl_driver.place(relx=0.1, rely=0.07)
lbl_server = Label(text='Server:')
lbl_server.place(relx=0.1, rely=0.17)
lbl_db = Label(text='Database:')
lbl_db.place(relx=0.1, rely=0.27)
lbl_uid = Label(text='UID:')
lbl_uid.place(relx=0.1, rely=0.37)
lbl_pwd = Label(text='Password:')
lbl_pwd.place(relx=0.1, rely=0.47)

combobox_driver = ttk.Combobox(values=drivers_list, state='readonly')
combobox_driver.place(relx=0.35, rely=0.07, width=200)
entry_server = Entry()
entry_server.place(relx=0.35, rely=0.17, width=200)
entry_db = Entry()
entry_db.place(relx=0.35, rely=0.27, width=200)
entry_uid = Entry()
entry_uid.place(relx=0.35, rely=0.37, width=200)
entry_pwd = Entry(show='*')
entry_pwd.place(relx=0.35, rely=0.47, width=200)

btn_connect = Button(window, text='Download SQL Log Data', command=btn_connect_func)
btn_connect.place(relx=0.17, rely=0.59, width=250)

frame_message = Frame(borderwidth=4, relief=RIDGE)

lbl_message = Label(frame_message)
lbl_message.pack(expand=True)


window.mainloop()
